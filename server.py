from datetime import datetime
import json
import os
from pathlib import Path
import shutil
import requests

from flask import Flask, send_file
from flask_restful import Resource, Api, reqparse
from flask_cors import CORS
import werkzeug
import pandas as pd
from openpyxl import load_workbook

import logging

from db_uploader import db_upload

import sys

import boto3
from botocore.exceptions import ClientError

from dataframe_utils import load_excel, remap_columns, save_as_excel


logger = logging.getLogger()
logger.setLevel(logging.INFO)

app = Flask(__name__)
api = Api(app)
CORS(app)

s3_client = boto3.client('s3')

# FILES_ROOT = Path(os.getenv("FILES_ROOT"))
# FILES_ROOT = Path(__file__).parent  # v.local
FILES_ROOT = Path("/tmp") # v.for lambda temp folder

CONFIG_PATH = Path(__file__).parent / "config.json"  #  v.local
with open(CONFIG_PATH, "r") as fp:
    CONFIG = json.load(fp)

bucket = os.environ['TARGET_BUCKET'] if os.environ.get('LAMBDA_TASK_ROOT') is not None else "mycanopybucket"

@app.route("/")
def hello():
    return "Canopy: Ad Astra Per Aspera! "+ ("We are offline" if os.environ.get('LAMBDA_TASK_ROOT') is None else "We are online")


class UploadFile(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument(
        "file",
        type=werkzeug.datastructures.FileStorage, 
        location='files'
    )
    def post(self):
        app.logger.debug("uploading file")

        args = self.parser.parse_args()
        file = args["file"]
        file.save(FILES_ROOT / "input.xlsx")
        print("Is input in tmp? " + str(os.path.isfile('/tmp/input.xlsx')), file=sys.stderr)
        # app.logger.error(print(os.path.isfile('/tmp/input.xlsx'))
       
        """
        Adding the S3 Upload
        """
        try:
            s3_client.upload_file(str(FILES_ROOT)+"/input.xlsx", bucket, "input.xlsx")
        # except ClientError as error:
        except:
            raise
        return self._sheetnames(file)


    def get(self):
        """
        Get the list of available sheet names
        """
        app.logger.debug("asking for sheet names")
        file_path = FILES_ROOT / "input.xlsx"
        if file_path.exists():
            return self._sheetnames(file_path)
        else:
            try:
              s3_client.download_file(bucket, "input.xlsx", str(FILES_ROOT)+ "/input.xlsx")
              return self._sheetnames(file_path)
            
            # except ClientError as error:
            except:
               print("No input file retrieved from S3", file=sys.stderr)
               return self._sheetnames()
            

    def _sheetnames(self, file=None):
        if file is not None:
            excel_file = pd.ExcelFile(file)
            sheet_names = list(excel_file.sheet_names)
        else:
            sheet_names = []
        return {"sheet_names": sheet_names}


api.add_resource(UploadFile, "/upload")


# For Bubble: bubble sends a link instead 
class UploadFileB(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument(
        "filepath",
         type=str
    )
    
    def post(self):
        app.logger.debug("uploading file")

        args = self.parser.parse_args()
        if args["filepath"] != "":
            url = 'https:'+args["filepath"]
            print(url, file=sys.stderr)
            r = requests.get(url, allow_redirects=True)        
            open(FILES_ROOT / "input.xlsx", 'wb').write(r.content)
        
            print("Is input in tmp? " + str(os.path.isfile('/tmp/input.xlsx')), file=sys.stderr)
            # app.logger.error(print(os.path.isfile('/tmp/input.xlsx'))
        
            """
            Adding the S3 Upload
            """
            try:
                s3_client.upload_file(str(FILES_ROOT)+"/input.xlsx", bucket, "input.xlsx")
            # except ClientError as error:
            except:
                raise

        return self._sheetnames(FILES_ROOT / "input.xlsx")


    def get(self):
        """
        Get the list of available sheet names
        """
        app.logger.debug("asking for sheet names")
        file_path = FILES_ROOT / "input.xlsx"
        if file_path.exists():
            return self._sheetnames(file_path)
        else:
            try:
              s3_client.download_file(bucket, "input.xlsx", str(FILES_ROOT)+ "/input.xlsx")
              return self._sheetnames(file_path)
            
            # except ClientError as error:
            except:
               print("No input file retrieved from S3", file=sys.stderr)
               return self._sheetnames()
            

    def _sheetnames(self, file=None):
        if file is not None:
            excel_file = pd.ExcelFile(file)
            sheet_names = list(excel_file.sheet_names)
        else:
            sheet_names = []
        return {"sheet_names": sheet_names}


api.add_resource(UploadFileB, "/uploadB")


class ProcessExcel(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument(
        "sheet_name",
        type=str,
    )

    def get(self):
        file_path = FILES_ROOT / "inter.feather"
        if file_path.exists():
            dataframe = pd.read_feather(file_path)
            return self._header(dataframe)
        return self._header()

    def post(self):
        """
        Read given excel sheet and process it. Return the resulting header
        """
        args = self.parser.parse_args()
        sheet_name = args["sheet_name"]
        """
        downloading input from s3 if file not in tmp 
        """
        if not (FILES_ROOT / "input.xlsx").exists():
            try:
              s3_client.download_file(bucket, "input.xlsx", str(FILES_ROOT)+ "/input.xlsx")
            except:
            # except ClientError as error:
               print("No input file retrieved from S3", file=sys.stderr)       

        dataframe = load_excel(
            FILES_ROOT / "input.xlsx", sheet_name=sheet_name
            ).astype(object)
        notnull_mask = dataframe.notnull()
        dataframe[notnull_mask] = dataframe[notnull_mask].astype(str)
        dataframe.to_feather(FILES_ROOT / "inter.feather")  # Not uploading the feather file now to S3 -> to check if needed

        if (FILES_ROOT / "db.json").exists():
            with open(FILES_ROOT / "db.json", "r") as fp:
                db = json.load(fp)
            saved_propositions = db.get("saved_propositions", {})
        else:
        # Download json from S3 if it exists 
            try:
                s3_client.download_file(bucket, "db.json", str(FILES_ROOT)+ "/db.json")
                with open(FILES_ROOT / "db.json", "r") as fp:
                    db = json.load(fp)
                saved_propositions = db.get("saved_propositions", {})
            except ClientError as e:
                # if e.response['Error']['Code'] == "404":
                    # The object does not exist.
                    saved_propositions = {}
                # else:
                #     # Something else has gone wrong.
                #     raise           

        mapping = {
            col: None for col in dataframe.columns
        }
        for (source, target) in saved_propositions.items():
            candidates = {
                key for key in CONFIG if key not in mapping.values()
            }
            if source in mapping and target in candidates:
                mapping[source] = target

        with open(FILES_ROOT / "db.json", "w") as fp:
            json.dump(
                {
                    "mapping": mapping,
                    "reference_date": None,
                    "saved_propositions": saved_propositions
                },
                fp
            )

        return self._header(dataframe)

    def _header(self, df=None):
        if df is not None:
            header = list(df.columns)
        else:
            header = None
        return {"header" : header}


api.add_resource(ProcessExcel, "/process")


@app.route("/download")
def download_file():
    file_path = FILES_ROOT / "output.xlsx"
    return send_file(file_path, as_attachment=True)


class MappingResource(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument("mapping", type=dict)
    parser.add_argument(
        "reference_date",
        required=False,
        nullable=True,
        type=datetime
    )

    def get(self):
        app.logger.debug("Reading current mapping")
        #adding a test for db.json and download it if not present
        if not (FILES_ROOT / "db.json").exists():
            s3_client.download_file(bucket, "db.json", str(FILES_ROOT)+ "/db.json")
        with open(FILES_ROOT / "db.json", "r") as fp:
            db = json.load(fp)
        return {
            "mapping": db["mapping"],
            "reference_date": db["reference_date"],
            "config": CONFIG
        }

    def _update_mapping(self, old_mapping, new_mapping):
        new_targets = set(new_mapping.values())
        return {
            **{
                key: value if value not in new_targets else None
                for (key, value) in old_mapping.items()
            },
            **new_mapping
        }

    def patch(self):
        app.logger.debug("Patching current mapping")        
        args = self.parser.parse_args()
        mapping = args["mapping"]
        reference_date = args["reference_date"]

        with open(FILES_ROOT / "db.json", "r") as fp:
            db = json.load(fp)
        new_mapping = self._update_mapping(db.get("mapping", {}), mapping)

        new_db = {
            **db,
            "mapping": new_mapping,
            "reference_date": reference_date
        }

        with open(FILES_ROOT / "db.json", "w") as fp:
            json.dump(new_db, fp)
        return {
            "mapping": new_db["mapping"], 
            "reference_date": new_db["reference_date"]
        }

    def post(self):
        app.logger.debug("validating current mapping")
        #not testing for feather file in tmp
        df = pd.read_feather(FILES_ROOT / "inter.feather")

        with open(FILES_ROOT / "db.json", "r") as fp:
            db = json.load(fp)
        mapping = db["mapping"]
        columns = {
            source: CONFIG[target]["name"]
            for (source, target) in mapping.items()
            if target is not None
        }
        # producing the output file with dropped columns and rows and remapped names
        output_df = remap_columns(df, columns, drop_unmapped=True)
    	
        # save csv for db_uploader function and save to db //OUTDATED
        # output_df.to_csv(FILES_ROOT / "outputdb.csv",index=False)

        # instead, directly send the dataframe to db_upload
        # output_df.to_csv(FILES_ROOT / "outputdb.csv",index=False)
        db_upload(output_df)
        
        template_path = Path(__file__).parent / "output_template.xlsx"
        if template_path.exists():
            shutil.copyfile(
                Path(__file__).parent / "output_template.xlsx",
                FILES_ROOT / "output.xlsx"
            )
        else:
            if (FILES_ROOT / "output.xlsx").exists():
                (FILES_ROOT / "output.xlsx").unlink()
        # saves as excel will transpose output_df
        save_as_excel(
            output_df, FILES_ROOT / "output.xlsx", sheet_name="input print",
            startrow=0
        )
        # added the upload of the output file to S3 
        s3_client.upload_file(str(FILES_ROOT)+"/output.xlsx", bucket, "output.xlsx")

        new_db = {
            **db,
            "saved_propositions": {
                **db.get("saved_propositions", {}), **mapping
            }
        }

        with open(FILES_ROOT / "db.json", "w") as fp:
            json.dump(new_db, fp)
        
        # added the upload of the db.json file to S3
        s3_client.upload_file(str(FILES_ROOT)+"/db.json", bucket, "db.json")

        return {}


api.add_resource(MappingResource, "/mapping")


#create a mapping for Bubble GET "/mappingB"

class MappingResourceB(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument("mapping", type=dict)
    parser.add_argument(
        "reference_date",
        required=False,
        nullable=True,
        type=datetime
    )

    def get(self):
        app.logger.debug("Reading current mapping")
        #adding a test for db.json and download it if not present
        if not (FILES_ROOT / "db.json").exists():
            s3_client.download_file(bucket, "db.json", str(FILES_ROOT)+ "/db.json")
        with open(FILES_ROOT / "db.json", "r") as fp:
            db = json.load(fp)
        
        mapping= db["mapping"]

        bubblejson = []
        for key, value in mapping.items():
            bubblejson.append({"row":key, "template":value})
        return {
            "mapping" : bubblejson
        }


api.add_resource(MappingResourceB, "/mappingB")


if __name__ == '__main__':
    HOST = os.getenv("HOST")
    app.logger.debug(f"Starting Server with host: {HOST}")
    app.run(debug=True, host=HOST)